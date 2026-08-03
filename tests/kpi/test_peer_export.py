"""Unit tests for peer_comparison.xlsx export — Sprint 3, Day 20."""

import sqlite3
import openpyxl

from src.analytics.peer import export_peer_comparison

DB_PATH = "data/nifty100.db"


def test_export_creates_exactly_11_sheets(tmp_path):
    output_path = str(tmp_path / "test_peer_comparison.xlsx")
    conn = sqlite3.connect(DB_PATH)
    export_peer_comparison(conn, output_path=output_path)
    conn.close()

    wb = openpyxl.load_workbook(output_path)
    assert len(wb.sheetnames) == 11


def test_export_it_services_sheet_has_correct_companies(tmp_path):
    output_path = str(tmp_path / "test_peer_comparison.xlsx")
    conn = sqlite3.connect(DB_PATH)
    export_peer_comparison(conn, output_path=output_path)
    conn.close()

    wb = openpyxl.load_workbook(output_path)
    ws = wb["IT Services"]

    company_ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row)]
    assert "TCS" in company_ids
    assert "INFY" in company_ids


def test_export_median_row_present(tmp_path):
    output_path = str(tmp_path / "test_peer_comparison.xlsx")
    conn = sqlite3.connect(DB_PATH)
    export_peer_comparison(conn, output_path=output_path)
    conn.close()

    wb = openpyxl.load_workbook(output_path)
    ws = wb["IT Services"]

    last_row_first_cell = ws.cell(row=ws.max_row, column=1).value
    assert last_row_first_cell == "MEDIAN"