"""Screener Filter Engine — Sprint 3, Day 15.

Loads screener_config.yaml and applies threshold filters to a combined
view of financial_ratios, market_cap, and sectors data.
"""

import sqlite3
import yaml
import pandas as pd

DB_PATH = "data/nifty100.db"
CONFIG_PATH = "config/screener_config.yaml"


def load_config():
    with open(CONFIG_PATH, "r") as f:
        return yaml.safe_load(f)


def get_latest_year_per_company(df, year_col="year"):
    """Reduce a multi-year DataFrame to just each company's latest year.

    Handles both text fiscal-year labels (e.g. '2024-03') and plain
    integer calendar years (e.g. 2024, as used in market_cap).
    """
    df = df.copy()

    if df[year_col].dtype == object:
        df["_year_sortable"] = df[year_col].str.replace("-", "").astype(int)
    else:
        df["_year_sortable"] = df[year_col]

    idx = df.groupby("company_id")["_year_sortable"].idxmax()
    return df.loc[idx].drop(columns=["_year_sortable"])


def build_screener_dataset(conn):
    """Combine financial_ratios + market_cap + sectors + companies + P&L,
    reduced to each company's latest available year.
    """
    ratios = pd.read_sql("SELECT * FROM financial_ratios", conn)
    ratios = get_latest_year_per_company(ratios)

    market_cap = pd.read_sql("SELECT * FROM market_cap", conn)
    market_cap = get_latest_year_per_company(market_cap)

    pl = pd.read_sql("SELECT company_id, year, sales, net_profit FROM profitandloss WHERE year != 'TTM'", conn)
    pl = get_latest_year_per_company(pl)

    sectors = pd.read_sql("SELECT company_id, broad_sector FROM sectors", conn)
    companies = pd.read_sql("SELECT id, company_name FROM companies", conn)

    merged = ratios.merge(market_cap.drop(columns=["year"]), on="company_id", how="left")
    merged = merged.merge(pl.drop(columns=["year"]), on="company_id", how="left")
    merged = merged.merge(sectors, on="company_id", how="left")
    merged = merged.merge(companies, left_on="company_id", right_on="id", how="left")

    return merged


def apply_filters(dataset, active_filters, config):
    """Apply a dict of {filter_name: threshold_value} to the dataset.

    active_filters example: {"roe_min": 15, "de_max": 1.0}
    """
    filters_config = config["filters"]
    result = dataset.copy()

    for filter_name, threshold in active_filters.items():
        if filter_name not in filters_config:
            continue

        rule = filters_config[filter_name]
        column = rule["column"]
        comparison = rule["comparison"]

        if rule.get("skip_for_sector"):
            skip_mask = result["broad_sector"] == rule["skip_for_sector"]
        else:
            skip_mask = pd.Series(False, index=result.index)

        if rule.get("treat_none_as_infinity"):
            values = result[column].fillna(float("inf"))
        else:
            values = result[column]

        if comparison == "min":
            passes = (values >= threshold) | skip_mask
        elif comparison == "exact":
            passes = (values == threshold) | skip_mask
        else:
            passes = (values <= threshold) | skip_mask

        result = result[passes.fillna(False)]
        
        if "composite_quality_score" in result.columns:
            result = result.sort_values("composite_quality_score", ascending=False)

    return result

def run_preset(dataset, preset_name, config):
    """Run one of the 5 simple threshold-based presets."""
    preset_filters = config["presets"][preset_name]
    return apply_filters(dataset, preset_filters, config)


def run_turnaround_watch(conn, dataset):
    """Turnaround Watch: Revenue CAGR 3yr > 10%, FCF positive in latest
    year, D/E declining year-over-year.
    """
    ratios_all_years = pd.read_sql("SELECT * FROM financial_ratios", conn)
    ratios_all_years = ratios_all_years[ratios_all_years["year"].str.endswith("-03")]

    results = []
    for company_id in dataset["company_id"].unique():
        company_history = ratios_all_years[ratios_all_years["company_id"] == company_id].sort_values("year")

        if len(company_history) < 2:
            continue

        latest = company_history.iloc[-1]
        previous = company_history.iloc[-2]

        if pd.isna(latest["debt_to_equity"]) or pd.isna(previous["debt_to_equity"]):
            continue
        de_declining = latest["debt_to_equity"] < previous["debt_to_equity"]

        fcf_positive = pd.notna(latest["free_cash_flow_cr"]) and latest["free_cash_flow_cr"] > 0

        rev_cagr_3yr = compute_revenue_cagr_3yr_for_screener(conn, company_id, latest["year"])
        cagr_ok = rev_cagr_3yr is not None and rev_cagr_3yr > 10

        if de_declining and fcf_positive and cagr_ok:
            results.append(company_id)

    return dataset[dataset["company_id"].isin(results)]

def compute_revenue_cagr_3yr_for_screener(conn, company_id, current_year):
    """Compute 3yr Revenue CAGR on demand for Turnaround Watch.

    Not stored in financial_ratios (Sprint 2 only stored 5yr CAGR),
    so we compute it directly here using the existing CAGR engine.
    """
    from src.analytics.cagr import compute_company_cagr

    pl = pd.read_sql(
        "SELECT company_id, year, sales FROM profitandloss WHERE company_id = ? AND year != 'TTM'",
        conn, params=(company_id,)
    )
    cagr, flag = compute_company_cagr(pl, "sales", current_year, 3)
    return cagr

def compute_sector_relative_composite_score(dataset):
    """Recompute composite_quality_score, but winsorise/normalise each
    metric WITHIN each broad_sector rather than across the whole universe.

    Uses the same weights as the Sprint 2 global score (Section 13):
    35% Profitability (ROE 15 + ROCE 10 + NPM 10) + 30% Cash Quality
    (FCF 15 + CFO/PAT 10 + FCF-positive-flag 5, simplified here to FCF
    only since CFO/PAT isn't stored per-row) + 20% Growth (Revenue CAGR
    10 + PAT CAGR 10) + 15% Leverage (D/E 10 + ICR 5).
    """
    from src.analytics.ratio_engine import winsorize_and_score

    df = dataset.copy()
    df["sector_composite_score"] = 0.0

    for sector, group in df.groupby("broad_sector"):
        idx = group.index

        roe_score = winsorize_and_score(group["return_on_equity_pct"].fillna(group["return_on_equity_pct"].median()))
        roce_score = winsorize_and_score(group["return_on_capital_employed_pct"].fillna(group["return_on_capital_employed_pct"].median()))
        npm_score = winsorize_and_score(group["net_profit_margin_pct"].fillna(group["net_profit_margin_pct"].median()))

        fcf_score = winsorize_and_score(group["free_cash_flow_cr"].fillna(group["free_cash_flow_cr"].median()))

        rev_cagr_score = winsorize_and_score(group["revenue_cagr_5yr"].fillna(group["revenue_cagr_5yr"].median()))
        pat_cagr_score = winsorize_and_score(group["pat_cagr_5yr"].fillna(group["pat_cagr_5yr"].median()))

        de_inverted = -group["debt_to_equity"].fillna(group["debt_to_equity"].median())
        de_score = winsorize_and_score(de_inverted)
        icr_filled = group["interest_coverage"].fillna(group["interest_coverage"].median())
        icr_score = winsorize_and_score(icr_filled)

        profitability = 0.15 * roe_score + 0.10 * roce_score + 0.10 * npm_score
        cash_quality = 0.30 * fcf_score
        growth = 0.10 * rev_cagr_score + 0.10 * pat_cagr_score
        leverage = 0.10 * de_score + 0.05 * icr_score

        sector_score = profitability + cash_quality + growth + leverage
        df.loc[idx, "sector_composite_score"] = sector_score.values

    return df

def export_screener_output(conn, config, output_path="output/screener_output.xlsx"):
    """Generate screener_output.xlsx — one sheet per preset, colour-coded
    cells (green = meets threshold, red = fails threshold).
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    dataset = build_screener_dataset(conn)
    dataset = compute_sector_relative_composite_score(dataset)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    display_cols = [
        "company_id", "company_name", "broad_sector",
        "return_on_equity_pct", "return_on_capital_employed_pct",
        "net_profit_margin_pct", "operating_profit_margin_pct",
        "debt_to_equity", "interest_coverage", "asset_turnover",
        "free_cash_flow_cr", "revenue_cagr_5yr", "pat_cagr_5yr",
        "eps_cagr_5yr", "pe_ratio", "pb_ratio", "dividend_yield_pct",
        "dividend_payout_ratio_pct", "sales", "net_profit",
        "composite_quality_score", "sector_composite_score",
    ]

    wb = Workbook()
    wb.remove(wb.active)

    for preset_name, preset_filters in config["presets"].items():
        result = run_preset(dataset, preset_name, config)
        result = result[[c for c in display_cols if c in result.columns]]

        ws = wb.create_sheet(title=preset_name[:31])

        ws.append(list(result.columns))

        for _, row in result.iterrows():
            ws.append(list(row))

        preset_filter_cols = {
            config["filters"][fname]["column"]: (config["filters"][fname]["comparison"], threshold)
            for fname, threshold in preset_filters.items()
            if fname in config["filters"]
        }

        header = list(result.columns)
        for row_idx in range(2, ws.max_row + 1):
            for col_idx, col_name in enumerate(header, start=1):
                if col_name in preset_filter_cols:
                    comparison, threshold = preset_filter_cols[col_name]
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    if value is None:
                        continue
                    if comparison == "min":
                        meets = value >= threshold
                    elif comparison == "exact":
                        meets = value == threshold
                    else:
                        meets = value <= threshold
                    cell.fill = green_fill if meets else red_fill

    # Turnaround Watch: special case, not YAML-driven (needs year-over-year comparison)
    turnaround_result = run_turnaround_watch(conn, dataset)
    turnaround_result = turnaround_result[[c for c in display_cols if c in turnaround_result.columns]]
    ws = wb.create_sheet(title="turnaround_watch")
    ws.append(list(turnaround_result.columns))
    for _, row in turnaround_result.iterrows():
        ws.append(list(row))
    # No per-cell colour-coding here since thresholds involve year-over-year
    # comparison logic, not simple static column thresholds.

    wb.save(output_path)
    return output_path