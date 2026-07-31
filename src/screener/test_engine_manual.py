"""Manual verification for the screener engine — Sprint 3, Day 15."""

import sqlite3
import pandas as pd

from src.screener.engine import build_screener_dataset, load_config
from src.screener.engine import apply_filters

DB_PATH = "data/nifty100.db"



if __name__ == "__main__":
    from src.screener.engine import export_screener_output

    conn = sqlite3.connect(DB_PATH)
    config = load_config()
    path = export_screener_output(conn, config)
    print("Saved:", path)
    conn.close()

    import openpyxl
    wb = openpyxl.load_workbook(path)
    print("Sheet names:", wb.sheetnames)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"{sheet_name}: {ws.max_row - 1} rows, {ws.max_column} columns")