"""Peer Percentile Ranking Engine — Sprint 3, Day 18.

Computes PERCENT_RANK for 10 metrics within each of 11 peer groups,
with D/E inverted (lower is better) per Section 18 of the project spec.
"""

import sqlite3
import pandas as pd

DB_PATH = "data/nifty100.db"

METRICS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "eps_cagr_5yr",
    "interest_coverage",
    "asset_turnover",
]

INVERTED_METRICS = {"debt_to_equity"}


def get_latest_year_per_company(df, year_col="year"):
    """Reduce a multi-year DataFrame to just each company's latest year."""
    df = df.copy()
    if df[year_col].dtype == object:
        df["_year_sortable"] = df[year_col].str.replace("-", "").astype(int)
    else:
        df["_year_sortable"] = df[year_col]
    idx = df.groupby("company_id")["_year_sortable"].idxmax()
    return df.loc[idx].drop(columns=["_year_sortable"])


def compute_peer_percentiles(conn):
    """Compute percentile ranks for all 10 metrics within each peer group.

    Returns a DataFrame ready to insert into peer_percentiles.
    Companies not in any peer group are simply absent from the result
    (handled gracefully, not an error).
    """
    ratios = pd.read_sql("SELECT * FROM financial_ratios", conn)
    ratios = get_latest_year_per_company(ratios)

    peer_groups = pd.read_sql("SELECT peer_group_name, company_id FROM peer_groups", conn)
    merged = peer_groups.merge(ratios, on="company_id", how="left")

    records = []
    for group_name in merged["peer_group_name"].unique():
        group_df = merged[merged["peer_group_name"] == group_name]

        for metric in METRICS:
            if metric not in group_df.columns:
                continue

            valid = group_df.dropna(subset=[metric])
            if len(valid) == 0:
                continue

            if metric in INVERTED_METRICS:
                ranks = 1 - valid[metric].rank(pct=True)
            else:
                ranks = valid[metric].rank(pct=True)

            for (_, row), pct_rank in zip(valid.iterrows(), ranks):
                records.append({
                    "company_id": row["company_id"],
                    "peer_group_name": group_name,
                    "metric": metric,
                    "value": row[metric],
                    "percentile_rank": pct_rank,
                    "year": row["year"],
                })

    columns = ["company_id", "peer_group_name", "metric", "value", "percentile_rank", "year"]
    return pd.DataFrame(records, columns=columns)

def write_peer_percentiles(conn, percentiles_df):
    """Write computed percentiles into the peer_percentiles table."""
    conn.execute("DELETE FROM peer_percentiles")
    conn.commit()
    percentiles_df.to_sql("peer_percentiles", conn, if_exists="append", index=False)
    
def export_peer_comparison(conn, output_path="output/peer_comparison.xlsx"):
    """Generate peer_comparison.xlsx — 11 sheets, one per peer group,
    with percentile colour-coding, benchmark highlighting, and a
    median summary row.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gold_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    ratios = pd.read_sql("SELECT * FROM financial_ratios", conn)
    ratios = get_latest_year_per_company(ratios)
    companies = pd.read_sql("SELECT id, company_name FROM companies", conn)
    peer_groups = pd.read_sql("SELECT peer_group_name, company_id, is_benchmark FROM peer_groups", conn)
    percentiles = compute_peer_percentiles(conn)

    wb = Workbook()
    wb.remove(wb.active)

    for group_name in peer_groups["peer_group_name"].unique():
        members = peer_groups[peer_groups["peer_group_name"] == group_name]
        group_ratios = members.merge(ratios, on="company_id", how="left")
        group_ratios = group_ratios.merge(companies, left_on="company_id", right_on="id", how="left")

        ws = wb.create_sheet(title=group_name[:31])

        metric_cols = [m for m in METRICS if m in group_ratios.columns]
        header = ["company_id", "company_name"] + metric_cols + [f"{m}_percentile" for m in metric_cols]
        ws.append(header)

        group_percentiles = percentiles[percentiles["peer_group_name"] == group_name]

        for _, row in group_ratios.iterrows():
            values = [row["company_id"], row["company_name"]]
            values += [row.get(m) for m in metric_cols]

            for m in metric_cols:
                pct_row = group_percentiles[
                    (group_percentiles["company_id"] == row["company_id"]) &
                    (group_percentiles["metric"] == m)
                ]
                values.append(pct_row["percentile_rank"].iloc[0] if len(pct_row) > 0 else None)

            ws.append(values)

        n_metric_cols = len(metric_cols)
        pct_col_start = 2 + n_metric_cols + 1

        for row_idx in range(2, ws.max_row + 1):
            for offset, m in enumerate(metric_cols):
                col_idx = pct_col_start + offset
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    continue
                if cell.value >= 0.75:
                    cell.fill = green_fill
                elif cell.value <= 0.25:
                    cell.fill = red_fill
                else:
                    cell.fill = yellow_fill

        for row_idx in range(2, ws.max_row + 1):
            company_id = ws.cell(row=row_idx, column=1).value
            is_bench = members[members["company_id"] == company_id]["is_benchmark"]
            if len(is_bench) > 0 and is_bench.iloc[0]:
                for col_idx in range(1, len(header) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = gold_fill

        median_row = ["MEDIAN", ""]
        for m in metric_cols:
            median_row.append(group_ratios[m].median())
        median_row += [None] * n_metric_cols
        ws.append(median_row)

    wb.save(output_path)
    return output_path