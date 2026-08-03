"""Manual verification for peer percentile rankings — Sprint 3, Day 18."""

import sqlite3

from src.analytics.peer import compute_peer_percentiles

DB_PATH = "data/nifty100.db"

if __name__ == "__main__":
    from src.analytics.peer import export_peer_comparison

    conn = sqlite3.connect(DB_PATH)
    path = export_peer_comparison(conn)
    conn.close()

    print("Saved:", path)

    import openpyxl
    wb = openpyxl.load_workbook(path)
    print("Sheet count:", len(wb.sheetnames))

    ws = wb["IT Services"]
    print()
    print(f"IT Services sheet: {ws.max_row} rows, {ws.max_column} columns")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        print(row[:4])